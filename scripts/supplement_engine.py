#!/usr/bin/env python3
"""
TruthTracer Supplement Engine — Fetch & Audit Supplementary Materials
=====================================================================
Detects data fabrication by cross-verifying supplementary data against
the main manuscript. Many fraudsters neglect to fabricate consistent SI.

Methods:
  1. Fetch SI links from DOI (CrossRef API)
  2. Download & parse SI (PDF, Excel, CSV)
  3. Extract tables & statistics from SI
  4. Cross-verify SI numbers vs main text claims
  5. SI completeness audit (are promised files present?)
  6. Excel sheet consistency (do formulas compute correctly?)

References:
  - COPE guidelines on supplementary materials
  - Bik EM et al. (2016). The prevalence of inappropriate image duplication. mBio
"""

import urllib.request, urllib.parse, json, os, re, tempfile, math
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from collections import defaultdict

# ═══════════════════════════════════════════════════════════
# 1. FETCH SUPPLEMENTARY MATERIAL LINKS
# ═══════════════════════════════════════════════════════════

def _get_json(url: str, timeout: int = 15) -> Optional[dict]:
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'TruthTracer/2.2 (mailto:researcher@example.com)',
            'Accept': 'application/json'
        })
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return json.loads(resp.read())
    except:
        return None

def fetch_si_from_doi(doi: str) -> dict:
    """
    Fetch supplementary material information from a paper's DOI.
    
    Uses CrossRef API to get:
    - Direct SI links
    - Resource links (data repositories, Zenodo, Figshare)
    - Journal SI pages
    
    Args:
        doi: Paper DOI (e.g., "10.1016/j.biortech.2023.123456")
    
    Returns:
        dict with SI links and metadata
    """
    # CrossRef work endpoint
    url = f"https://api.crossref.org/works/{urllib.parse.quote(doi)}"
    data = _get_json(url)
    
    if not data or 'message' not in data:
        return {"doi": doi, "found": False, "error": "DOI not found in CrossRef"}
    
    msg = data['message']
    si_links = []
    
    # Check for direct SI links
    for link in msg.get('link', []):
        content_type = link.get('content-type', '')
        intended_app = link.get('intended-application', '')
        url_link = link.get('URL', '')
        
        if 'supplement' in content_type.lower() or 'supplement' in intended_app.lower():
            si_links.append({
                'url': url_link,
                'type': link.get('content-type', 'unknown'),
                'description': link.get('content-version', '')
            })
    
    # Check resource links (data repositories)
    resource = msg.get('resource', {})
    primary = resource.get('primary', {})
    if primary.get('URL'):
        si_links.append({
            'url': primary['URL'],
            'type': 'resource',
            'description': 'Publisher SI page'
        })
    
    return {
        "doi": doi,
        "found": True,
        "title": msg.get('title', ['Unknown'])[0] if msg.get('title') else 'Unknown',
        "publisher": msg.get('publisher', 'Unknown'),
        "journal": msg.get('container-title', ['Unknown'])[0] if msg.get('container-title') else 'Unknown',
        "n_si_links": len(si_links),
        "si_links": si_links,
        "has_data_repository": any('repository' in str(l).lower() or 'zenodo' in str(l).lower() or 'figshare' in str(l).lower() for l in si_links)
    }


# ═══════════════════════════════════════════════════════════
# 2. DOWNLOAD & PARSE SUPPLEMENTARY FILES
# ═══════════════════════════════════════════════════════════

def download_si_file(url: str, output_dir: str = None) -> Optional[str]:
    """
    Download a supplementary file (PDF, Excel, CSV).
    
    Returns path to downloaded file, or None if failed.
    """
    if not output_dir:
        output_dir = tempfile.mkdtemp(prefix="truthtracer_si_")
    
    # Determine filename from URL
    filename = url.split('/')[-1].split('?')[0]
    if not filename or '.' not in filename:
        filename = f"si_download_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    filepath = os.path.join(output_dir, filename)
    
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'TruthTracer/2.2 (mailto:researcher@example.com)'
        })
        with urllib.request.urlopen(req, timeout=30) as resp:
            content = resp.read()
        
        with open(filepath, 'wb') as f:
            f.write(content)
        
        return filepath
    except Exception as e:
        return None


def parse_si_excel(filepath: str) -> dict:
    """
    Parse an Excel supplementary file and extract all numeric tables.
    
    Returns dict of sheet_name -> list of rows with numeric data.
    """
    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed", "file": filepath}
    
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return {"error": f"Cannot open Excel: {e}", "file": filepath}
    
    sheets_data = {}
    all_numbers = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        sheet_numbers = []
        
        for row in ws.iter_rows(values_only=True):
            # Extract only numeric cells
            numeric_row = [cell for cell in row if isinstance(cell, (int, float))]
            if numeric_row:
                rows.append(numeric_row)
                sheet_numbers.extend(numeric_row)
        
        if rows:
            sheets_data[sheet_name] = {
                "n_rows": len(rows),
                "n_numeric_cells": len(sheet_numbers),
                "sample_row": rows[0][:10] if rows else [],
                "all_values": sheet_numbers[:500]  # First 500 for analysis
            }
            all_numbers.extend(sheet_numbers)
    
    wb.close()
    
    return {
        "file": filepath,
        "n_sheets": len(sheets_data),
        "sheets": sheets_data,
        "total_numbers": len(all_numbers),
        "all_values": all_numbers[:1000]
    }


def parse_si_pdf(filepath: str) -> dict:
    """
    Parse a PDF supplementary file and extract text + embedded tables.
    """
    try:
        import fitz  # pymupdf
        doc = fitz.open(filepath)
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
    except ImportError:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            text = ""
            for page in reader.pages:
                t = page.extract_text()
                if t: text += t + "\n"
        except:
            return {"error": "No PDF library available", "file": filepath}
    
    # Extract all numbers from text
    numbers = re.findall(r'(?<!\d)(\d+\.?\d*)(?!\d)', text)
    numeric_values = [float(n) for n in numbers if 0.01 < float(n) < 100000]
    
    # Extract table-like structures (lines with 3+ numbers)
    lines = text.split('\n')
    table_lines = [l for l in lines if len(re.findall(r'\d+\.?\d*', l)) >= 3]
    
    # Find means ± SD patterns
    pm_patterns = re.findall(r'(\d+\.?\d*)\s*[±±]\s*(\d+\.?\d*)', text)
    
    return {
        "file": filepath,
        "text_length": len(text),
        "n_numeric_values": len(numeric_values),
        "n_table_lines": len(table_lines),
        "n_mean_sd_pairs": len(pm_patterns),
        "all_values": numeric_values[:500],
        "mean_sd_pairs": [(float(m), float(s)) for m, s in pm_patterns[:50]],
        "sample_table_lines": table_lines[:10]
    }


def parse_si_csv(filepath: str) -> dict:
    """Parse a CSV supplementary file."""
    import csv
    all_numbers = []
    rows = []
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                nums = []
                for cell in row:
                    try:
                        nums.append(float(cell))
                        all_numbers.append(float(cell))
                    except:
                        pass
                if nums:
                    rows.append(nums)
    except:
        return {"error": "Cannot parse CSV", "file": filepath}
    
    return {
        "file": filepath,
        "n_rows": len(rows),
        "n_numeric_cells": len(all_numbers),
        "all_values": all_numbers[:500],
        "sample_row": rows[0][:10] if rows else []
    }


# ═══════════════════════════════════════════════════════════
# 3. CROSS-VERIFY SI vs MAIN TEXT
# ═══════════════════════════════════════════════════════════

def cross_verify_numbers(main_text_numbers: List[float], 
                          si_numbers: List[float],
                          tolerance: float = 0.01) -> dict:
    """
    Check if numbers in the supplementary material match those in the main text.
    
    A common fabrication tell: numbers that appear in the main text but are
    absent from or inconsistent with the supplementary data tables.
    
    Args:
        main_text_numbers: Numbers extracted from the main manuscript
        si_numbers: Numbers extracted from supplementary files
        tolerance: Relative tolerance for matching
    
    Returns:
        dict with match statistics
    """
    if not main_text_numbers or not si_numbers:
        return {"verdict": "SKIP", "note": "Insufficient data for comparison"}
    
    si_set = set(round(n, 4) for n in si_numbers)
    
    matches = 0
    mismatches = []
    missing_from_si = []
    
    for n in main_text_numbers[:100]:
        rounded = round(n, 4)
        # Check if this number or a very close match exists in SI
        found = False
        for si_n in si_numbers[:500]:
            if si_n != 0 and abs(rounded - round(si_n, 4)) / max(abs(si_n), 0.001) < tolerance:
                found = True
                break
        if found:
            matches += 1
        else:
            missing_from_si.append(n)
    
    match_rate = matches / max(len(main_text_numbers[:100]), 1) * 100
    
    issues = []
    if match_rate < 50:
        issues.append(f"Only {match_rate:.0f}% of main-text numbers found in SI — possible data inconsistency")
    elif match_rate < 80:
        issues.append(f"{match_rate:.0f}% match rate — moderate SI-main text discrepancy")
    
    return {
        "method": "SI-Main Text Cross-Verification",
        "main_text_numbers": len(main_text_numbers[:100]),
        "si_numbers": len(si_numbers[:500]),
        "matches": matches,
        "match_rate_pct": round(match_rate, 1),
        "missing_count": len(missing_from_si),
        "missing_examples": missing_from_si[:10],
        "issues": issues,
        "verdict": "RED FLAG" if match_rate < 50 else \
                   "WARNING" if match_rate < 80 else "PASS"
    }


# ═══════════════════════════════════════════════════════════
# 4. SI COMPLETENESS AUDIT
# ═══════════════════════════════════════════════════════════

def audit_si_completeness(main_text: str, si_info: dict, 
                           si_parsed_data: dict = None) -> dict:
    """
    Check whether all supplementary materials promised in the main text
    are actually present and accessible.
    
    Flags:
    - "See Supplementary Table S1" but S1 is missing from SI
    - "Data available at [URL]" but URL is broken
    - "Supplementary Figures S1-S10" but only S1-S5 are present
    
    Args:
        main_text: Full text of the main manuscript
        si_info: Output from fetch_si_from_doi()
        si_parsed_data: Output from parse_si_excel/pdf/csv()
    
    Returns:
        dict with completeness audit
    """
    # Find all SI references in main text
    si_refs = re.findall(
        r'(?:Supplementary|Supplemental|Supporting)\s+(?:Table|Figure|Material|Data|File|Note|Text)\s+S?(\d+)',
        main_text, re.IGNORECASE
    )
    si_refs_numeric = [int(r) for r in si_refs if r.isdigit()]
    
    issues = []
    
    # Check: Are SI files accessible?
    if si_info.get('found') and si_info.get('n_si_links', 0) == 0:
        issues.append("Paper mentions supplementary materials but no SI links found via CrossRef")
    
    # Check: Do promised SI tables/figures exist?
    if si_refs_numeric and si_parsed_data:
        max_promised = max(si_refs_numeric) if si_refs_numeric else 0
        n_si_files = 0
        if isinstance(si_parsed_data, dict):
            for k, v in si_parsed_data.items():
                if isinstance(v, dict) and 'n_sheets' in v:
                    n_si_files += v['n_sheets']
        
        if max_promised > 0 and n_si_files > 0 and n_si_files < max_promised * 0.5:
            issues.append(
                f"Main text references up to S{max_promised}, "
                f"but only {n_si_files} SI tables/sheets found"
            )
    
    # Check data availability statements
    data_statement = re.search(
        r'Data (?:availability|accessibility|deposition).*?(?:\n\n|\n\s*\n|\.\s+[A-Z])',
        main_text, re.DOTALL | re.IGNORECASE
    )
    
    has_data_statement = bool(data_statement)
    data_urls = re.findall(r'https?://[^\s\)]+', data_statement.group(0)) if data_statement else []
    
    if not has_data_statement:
        issues.append("No data availability statement found")
    
    # Verify data URLs
    broken_urls = []
    for url in data_urls[:5]:
        try:
            req = urllib.request.Request(url, headers={'User-Agent': 'TruthTracer/2.2'})
            with urllib.request.urlopen(req, timeout=10) as resp:
                pass  # URL is accessible
        except:
            broken_urls.append(url)
    
    if broken_urls:
        issues.append(f"{len(broken_urls)} data availability URLs are inaccessible")
    
    return {
        "method": "SI Completeness Audit",
        "has_data_statement": has_data_statement,
        "n_si_refs_in_text": len(si_refs_numeric),
        "max_si_ref": max(si_refs_numeric) if si_refs_numeric else 0,
        "n_data_urls": len(data_urls),
        "n_broken_urls": len(broken_urls),
        "broken_urls": broken_urls[:5],
        "issues": issues,
        "verdict": "RED FLAG" if len(issues) >= 2 else \
                   "WARNING" if issues else "PASS"
    }


# ═══════════════════════════════════════════════════════════
# 5. FULL SI AUDIT PIPELINE
# ═══════════════════════════════════════════════════════════

def full_si_audit(doi: str, main_text: str = "", 
                   download: bool = False,
                   output_dir: str = None) -> dict:
    """
    Run the complete supplementary material audit pipeline.
    
    1. Fetch SI links from DOI
    2. Download SI files (if download=True)
    3. Parse SI content
    4. Cross-verify with main text
    5. Audit SI completeness
    
    Args:
        doi: Paper DOI
        main_text: Full text of the main manuscript (optional but recommended)
        download: Whether to download and parse SI files
        output_dir: Where to save downloaded files
    
    Returns:
        Complete SI audit report
    """
    results = {}
    
    # Step 1: Fetch SI info
    print("  [1/4] Fetching SI links from DOI...")
    si_info = fetch_si_from_doi(doi)
    results['si_info'] = si_info
    
    if not si_info.get('found'):
        return {"error": "DOI not found", "results": results}
    
    print(f"  Found {si_info.get('n_si_links', 0)} SI links")
    
    # Step 2: Download & parse
    si_parsed = {}
    if download and si_info.get('si_links'):
        print(f"  [2/4] Downloading SI files...")
        for i, link in enumerate(si_info['si_links']):
            url = link.get('url', '')
            if not url:
                continue
            print(f"    Downloading: {url[:80]}...")
            filepath = download_si_file(url, output_dir)
            if not filepath:
                continue
            
            # Parse based on file type
            ext = os.path.splitext(filepath)[1].lower()
            if ext in ['.xlsx', '.xls']:
                parsed = parse_si_excel(filepath)
            elif ext == '.pdf':
                parsed = parse_si_pdf(filepath)
            elif ext == '.csv':
                parsed = parse_si_csv(filepath)
            else:
                parsed = {"file": filepath, "type": "unknown"}
            
            si_parsed[f"si_file_{i+1}"] = parsed
            results['si_parsed'] = si_parsed
    
    # Step 3: Cross-verify
    if main_text and si_parsed:
        print(f"  [3/4] Cross-verifying SI vs main text...")
        # Extract numbers from main text
        main_numbers = [float(n) for n in re.findall(r'(?<!\d)(\d+\.?\d*)(?!\d)', main_text)
                       if 0.01 < float(n) < 100000]
        
        # Collect SI numbers
        si_all_numbers = []
        for key, parsed in si_parsed.items():
            if isinstance(parsed, dict) and 'all_values' in parsed:
                si_all_numbers.extend(parsed['all_values'])
        
        results['cross_verify'] = cross_verify_numbers(main_numbers, si_all_numbers)
    
    # Step 4: Completeness audit
    if main_text:
        print(f"  [4/4] Auditing SI completeness...")
        results['completeness'] = audit_si_completeness(main_text, si_info, si_parsed)
    
    # Summary
    red = sum(1 for r in results.values() 
             if isinstance(r, dict) and 'RED FLAG' in r.get('verdict', ''))
    warn = sum(1 for r in results.values() 
              if isinstance(r, dict) and 'WARNING' in r.get('verdict', ''))
    
    results['summary'] = {
        "si_available": si_info.get('n_si_links', 0) > 0,
        "si_downloaded": len(si_parsed),
        "red_flags": red,
        "warnings": warn,
        "overall": "RED FLAG" if red >= 2 else "WARNING" if red >= 1 or warn >= 2 else "PASS"
    }
    
    return {"doi": doi, "results": results}


# ═══════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════

if __name__ == '__main__':
    import sys
    if len(sys.argv) < 2:
        print("TruthTracer Supplement Engine")
        print("Usage:")
        print("  python supplement_engine.py fetch <doi>              — Get SI links")
        print("  python supplement_engine.py audit <doi> <text_file>  — Full SI audit")
        print("  python supplement_engine.py parse <excel_file>       — Parse SI Excel")
        sys.exit(1)
    
    cmd = sys.argv[1]
    
    if cmd == 'fetch':
        result = fetch_si_from_doi(sys.argv[2])
        print(json.dumps(result, indent=2, ensure_ascii=False))
    
    elif cmd == 'audit':
        doi = sys.argv[2]
        main_text = ""
        if len(sys.argv) > 3:
            with open(sys.argv[3], 'r', encoding='utf-8') as f:
                main_text = f.read()
        result = full_si_audit(doi, main_text, download=True)
        print(json.dumps(result, indent=2, ensure_ascii=False))
    
    elif cmd == 'parse':
        filepath = sys.argv[2]
        ext = os.path.splitext(filepath)[1].lower()
        if ext in ['.xlsx', '.xls']:
            result = parse_si_excel(filepath)
        elif ext == '.pdf':
            result = parse_si_pdf(filepath)
        elif ext == '.csv':
            result = parse_si_csv(filepath)
        else:
            result = {"error": f"Unsupported format: {ext}"}
        print(json.dumps(result, indent=2, ensure_ascii=False))
